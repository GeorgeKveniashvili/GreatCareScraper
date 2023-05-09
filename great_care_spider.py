import openpyxl
from scrapy.spiders import CrawlSpider, Rule
from scrapy.linkextractors import LinkExtractor

# Website URL to crawl
GREAT_CARE_URL = 'https://www.iwantgreatcare.org/search'
# Regex for subdomains that are scraped
ALLOWED_EXTRACTOR = ['/doctors/', '/optometrists/', '/nurses/', '/dentists/', '/physiotherapists/', '/dietitians/', '/occupationaltherapists/']

# XPath for person's full name
TITLE_XPATH = '//*[@id="entity-name-score-container"]/h1'
# Known XPathes for "Specialises in" field
SPECIALISES_XPATHES = ['//*[@id="specialies-container"]/div/ul', '//*[@id="information"]/div[2]/div[2]/div/ul', 
                       '//*[@id="information"]/div[1]/div[2]/div/ul']
# Known XPathes for "Works at" field
WORKS_XPATHES = ['//*[@id="works-at-container"]/div/ul', '//*[@id="information"]/div[2]/div[1]/div/ul', 
                 '//*[@id="information"]/div[1]/div[1]/div/ul', '/html/body/div[1]/main/div/span/div[2]/div/div[2]/div[1]/ul',
                 '/html/body/div[1]/main/div/span/div[2]/div/div[2]/div[2]/div[1]/div/ul']
# Known XPathes for "Profile" field
PROFILE_XPATHES = ['/html/body/div[1]/main/div/span/div[2]/div/div[2]/div[4]/div/div/p', 
                   '/html/body/div[1]/main/div/span/div[2]/div/div[2]/div[2]/div/div/p']

# Excel document path on local machine
FILE_PATH = "DoctorInfo.xlsx"
# Starting row index of the Excel document
FILE_INDEX = 2

# Main spider that is being run
class GreatCareSpider(CrawlSpider):
    # Website info
    name = 'greatcarespider'
    allowed_domains = ['iwantgreatcare.org']
    start_urls = [GREAT_CARE_URL]

    # Rules to crawl only required subdomains
    rules = (
        # Extract and follow all links!
        Rule(LinkExtractor(allow=ALLOWED_EXTRACTOR), callback='parse_item', follow=True),
    )

    # Method which is being executed if the rules match
    def parse_item(self, response):
        global FILE_INDEX
        
        # Get each title and scrape the info
        for title in response.xpath(TITLE_XPATH):
            # Split full name
            full_name_formated = split_name(title.css('::text').get())
            name = full_name_formated[0]
            surname = full_name_formated[1]

            # Extraxt "Specialises in", "Profile", and "Works at" fields
            specialises = extract_multiple_selectors(response, SPECIALISES_XPATHES, 'li::text', surname)
            profile = extract_multiple_selectors(response, PROFILE_XPATHES, '::text', surname)
            works = extract_multiple_selectors(response, WORKS_XPATHES, 'a::text', surname)

            # Save the data to the Excel document
            write_file(FILE_INDEX, name, surname, specialises, profile, works, response.url)
            FILE_INDEX += 1
        self.log('crawling'.format(response.url))

# Try all known XPathes and detect which is present on the current webpage
def extract_multiple_selectors(response, xpathes, text_selector, name):
    for xpath in xpathes:
        # Get object with the XPath and all inside texts
        selector = response.xpath(xpath)
        text_list = selector.css(text_selector).getall()
        # Check if the selected result is viable
        if len(text_list) > 0:
            print("The selector is good: "+ str(selector)+"name: "+name)
            break
        print("The selector is None: "+ str(selector)+"name: "+name)
    # Append each item in the list of texts to a single string
    text = ""
    for item in text_list:
        text += str(item)
    return text

# Open, write, and save the Excel document
def write_file(index, name, surname, specialises, profile, works, domain):
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.get_sheet_by_name('Sheet')

    ws.cell(row=index, column=1).value = name
    ws.cell(row=index, column=2).value = surname
    ws.cell(row=index, column=3).value = specialises
    ws.cell(row=index, column=4).value = profile
    ws.cell(row=index, column=5).value = works
    ws.cell(row=index, column=6).value = domain

    wb.save(FILE_PATH)
    wb.close()

# Split full name into name and surname
def split_name(full_name):
    # Check if the full name consists of more than 2 words
    if(len(full_name.split(' ')) > 2):
        # Take only first 2 words and set is as a name
        # Take the rest of the words and set it as a surname
        name = ' '.join(full_name.split(' ', 2)[:2])
        surname = ' '.join(full_name.split(' ', 2)[2:])
        return [name, surname]
    # If the full name is only 2 words
    # Split it in half and set name, surname accordingly
    name = full_name.split(' ', 1)[0]
    surname = full_name.split(' ', 1)[1]
    return [name, surname]